from sqlalchemy.orm import Session
from typing import List, Optional, Union, Any
import os
import io
import re
import json
import pandas as pd
import numpy as np
from datetime import datetime, date
from backend.models.dataset import Dataset
from backend.core.exceptions import NoActiveDatasetException
from backend.core.logging import logger

class DatasetService:
    def __init__(self, db: Session):
        self.db = db

    def get_active_dataset_id(self) -> Optional[int]:
        """
        Resolves the currently active dataset ID.
        """
        active_row = self.db.query(Dataset.id).filter(Dataset.is_active == True).first()
        return active_row[0] if active_row else None

    def get_active_dataset_id_or_raise(self) -> int:
        """
        Resolves the currently active dataset ID, or raises NoActiveDatasetException if none is active.
        """
        active_id = self.get_active_dataset_id()
        if active_id is None:
            raise NoActiveDatasetException()
        return active_id

    def get_active_dataset(self) -> Optional[Dataset]:
        """
        Gets the currently active Dataset object.
        """
        return self.db.query(Dataset).filter(Dataset.is_active == True).first()

    def list_datasets(self) -> List[Dataset]:
        """
        Lists all datasets in the registry sorted by creation time.
        """
        return self.db.query(Dataset).order_by(Dataset.created_at.desc()).all()

    def activate_dataset(self, dataset_id: int) -> Dataset:
        """
        Sets the specified dataset as active. Enforces the maximum active datasets limit,
        deactivating the oldest active dataset(s) if necessary.
        """
        from backend.models.dataset import DatasetConfig
        from datetime import datetime

        dataset = self.db.query(Dataset).filter(Dataset.id == dataset_id).first()
        if not dataset:
            raise ValueError(f"Dataset with ID {dataset_id} not found.")

        if dataset.status == "Archived":
            raise ValueError("Cannot activate an archived dataset.")

        if dataset.status != "Ready":
            raise ValueError("Cannot activate a dataset that is not in Ready state.")

        if dataset.is_active:
            return dataset

        # Load config
        config = self.db.query(DatasetConfig).first()
        max_active = config.max_active_datasets if config else "1"

        if max_active != "All":
            limit = int(max_active)
            # Find currently active datasets ordered by upload_time ascending (oldest first)
            active_datasets = self.db.query(Dataset).filter(Dataset.is_active == True).order_by(Dataset.upload_time.asc()).all()
            if len(active_datasets) >= limit:
                # Deactivate oldest active datasets to make space
                num_to_deactivate = len(active_datasets) - limit + 1
                for i in range(num_to_deactivate):
                    old_ds = active_datasets[i]
                    old_ds.is_active = False
                    logger.info(f"[Dataset Activation Change] Dataset ID: {old_ds.id}, Timestamp: {datetime.utcnow().isoformat()}, Previous Status: Active, New Status: Inactive")

        # Activate this dataset
        dataset.is_active = True
        logger.info(f"[Dataset Activation Change] Dataset ID: {dataset.id}, Timestamp: {datetime.utcnow().isoformat()}, Previous Status: Inactive, New Status: Active")

        self.db.commit()
        return dataset

    def deactivate_dataset(self, dataset_id: int) -> Dataset:
        """
        Sets the specified dataset as inactive.
        """
        from datetime import datetime
        dataset = self.db.query(Dataset).filter(Dataset.id == dataset_id).first()
        if not dataset:
            raise ValueError(f"Dataset with ID {dataset_id} not found.")

        if not dataset.is_active:
            return dataset

        # Deactivate
        dataset.is_active = False
        logger.info(f"[Dataset Activation Change] Dataset ID: {dataset.id}, Timestamp: {datetime.utcnow().isoformat()}, Previous Status: Active, New Status: Inactive")
        self.db.commit()
        return dataset

    def get_active_datasets(self) -> List[Dataset]:
        """
        Gets all currently active Dataset objects.
        """
        return self.db.query(Dataset).filter(Dataset.is_active == True).all()

    def get_active_dataset_ids(self) -> List[int]:
        """
        Gets the IDs of all currently active datasets.
        """
        active_ids = self.db.query(Dataset.id).filter(Dataset.is_active == True).all()
        return [row[0] for row in active_ids]

    def get_active_datasets_metadata(self) -> List[dict]:
        """
        Returns basic metadata dicts for all active datasets.
        """
        active_ds = self.get_active_datasets()
        return [
            {
                "id": ds.id,
                "name": ds.name,
                "display_name": ds.display_name,
                "original_filename": ds.original_filename,
                "row_count": ds.row_count,
                "column_count": ds.column_count,
                "file_size": ds.file_size
            }
            for ds in active_ds
        ]

    def get_dataset_config(self):
        """
        Retrieves the dataset settings configuration, creating the default one if missing.
        """
        from backend.models.dataset import DatasetConfig
        config = self.db.query(DatasetConfig).first()
        if not config:
            config = DatasetConfig(max_active_datasets="1")
            self.db.add(config)
            self.db.commit()
            self.db.refresh(config)
        return config

    def update_dataset_config(self, max_active: str):
        """
        Updates the dataset settings configuration and enforces active count limits.
        """
        from backend.models.dataset import DatasetConfig
        from datetime import datetime
        
        if max_active not in ["1", "2", "3", "All"]:
            raise ValueError("Invalid maximum active count. Options are '1', '2', '3', or 'All'.")

        config = self.get_dataset_config()
        config.max_active_datasets = max_active
        self.db.commit()

        # If the new limit is a number, we may need to deactivate some datasets
        if max_active != "All":
            limit = int(max_active)
            active_datasets = self.db.query(Dataset).filter(Dataset.is_active == True).order_by(Dataset.upload_time.asc()).all()
            if len(active_datasets) > limit:
                # Deactivate the oldest active datasets
                num_to_deactivate = len(active_datasets) - limit
                for i in range(num_to_deactivate):
                    old_ds = active_datasets[i]
                    old_ds.is_active = False
                    logger.info(f"[Dataset Activation Change] Dataset ID: {old_ds.id}, Timestamp: {datetime.utcnow().isoformat()}, Previous Status: Active, New Status: Inactive")
                self.db.commit()

        return config


    def delete_dataset(self, dataset_id: int) -> bool:
        """
        Soft-deletes (archives) the dataset from the registry. Sets status to 'Archived'
        and is_active to False.
        """
        dataset = self.db.query(Dataset).filter(Dataset.id == dataset_id).first()
        if not dataset:
            raise ValueError(f"Dataset with ID {dataset_id} not found.")

        # If active, automatically fallback to another ready dataset
        if dataset.is_active:
            fallback_ds = self.db.query(Dataset).filter(
                Dataset.id != dataset_id,
                Dataset.status == "Ready"
            ).first()
            if fallback_ds:
                fallback_ds.is_active = True
                logger.info(f"Automatically fell back active context to dataset ID {fallback_ds.id} because dataset ID {dataset_id} is being archived.")
            dataset.is_active = False

        # Soft delete: set status to Archived
        dataset.status = "Archived"
        self.db.commit()

        logger.info(f"Dataset ID {dataset_id} soft-deleted/archived successfully.")
        return True


    def delete_dataset_permanent(self, dataset_id: int) -> bool:
        """
        Hard-deletes the dataset and all its associated child rows from the database.
        Also deletes the underlying uploaded file from disk.
        """
        import os
        dataset = self.db.query(Dataset).filter(Dataset.id == dataset_id).first()
        if not dataset:
            raise ValueError(f"Dataset with ID {dataset_id} not found.")

        # If active, automatically fallback to another ready dataset
        if dataset.is_active:
            fallback_ds = self.db.query(Dataset).filter(
                Dataset.id != dataset_id,
                Dataset.status == "Ready"
            ).first()
            if fallback_ds:
                fallback_ds.is_active = True
                logger.info(f"Automatically fell back active context to dataset ID {fallback_ds.id} because dataset ID {dataset_id} is being permanently deleted.")
            dataset.is_active = False


        # Import models locally to avoid circular dependencies
        from backend.models.fir_case import CaseMaster, Inv_OccuranceTime
        from backend.models.fir_people import ComplainantDetails, FIRVictim, Accused
        from backend.models.fir_proceedings import ArrestSurrender, InvArrestSurrenderAccused, ChargesheetDetails
        from backend.models.fir_law import ActSectionAssociation
        from backend.models.crime import CrimeEvent
        from backend.models.criminal import Criminal
        from backend.models.victim import Victim
        from backend.models.crime_participation import CrimeParticipation

        case_ids = [c.id for c in dataset.cases]
        if case_ids:
            # 1. inv_arrestsurrenderaccused
            as_ids = [row[0] for row in self.db.query(ArrestSurrender.id).filter(ArrestSurrender.CaseMasterID.in_(case_ids)).all()]
            if as_ids:
                self.db.query(InvArrestSurrenderAccused).filter(InvArrestSurrenderAccused.ArrestSurrenderID.in_(as_ids)).delete(synchronize_session=False)
            
            # 2. arrest_surrender
            self.db.query(ArrestSurrender).filter(ArrestSurrender.CaseMasterID.in_(case_ids)).delete(synchronize_session=False)
            
            # 3. chargesheet_details
            self.db.query(ChargesheetDetails).filter(ChargesheetDetails.CaseMasterID.in_(case_ids)).delete(synchronize_session=False)
            
            # 4. complainant_details
            self.db.query(ComplainantDetails).filter(ComplainantDetails.CaseMasterID.in_(case_ids)).delete(synchronize_session=False)
            
            # 5. victim (FIRVictim)
            self.db.query(FIRVictim).filter(FIRVictim.CaseMasterID.in_(case_ids)).delete(synchronize_session=False)
            
            # 6. accused
            self.db.query(Accused).filter(Accused.CaseMasterID.in_(case_ids)).delete(synchronize_session=False)
            
            # 7. act_section_association
            self.db.query(ActSectionAssociation).filter(ActSectionAssociation.CaseMasterID.in_(case_ids)).delete(synchronize_session=False)
            
            # 8. inv_occurance_time
            self.db.query(Inv_OccuranceTime).filter(Inv_OccuranceTime.CaseMasterID.in_(case_ids)).delete(synchronize_session=False)

            # 9. Delete cases
            self.db.query(CaseMaster).filter(CaseMaster.id.in_(case_ids)).delete(synchronize_session=False)

        # Legacy data tables
        self.db.query(CrimeParticipation).filter(CrimeParticipation.dataset_id == dataset_id).delete(synchronize_session=False)
        self.db.query(CrimeEvent).filter(CrimeEvent.dataset_id == dataset_id).delete(synchronize_session=False)
        self.db.query(Criminal).filter(Criminal.dataset_id == dataset_id).delete(synchronize_session=False)
        self.db.query(Victim).filter(Victim.dataset_id == dataset_id).delete(synchronize_session=False)

        # Delete from disk if storage_path exists
        if dataset.storage_path and os.path.exists(dataset.storage_path):
            try:
                os.remove(dataset.storage_path)
            except Exception as io_err:
                logger.error(f"Error removing dataset file {dataset.storage_path}: {io_err}")

        # Delete dataset registry record
        self.db.delete(dataset)
        self.db.commit()
        return True

    def import_dataset(
        self,
        display_name: str,
        description: Optional[str],
        file_name: str,
        file_bytes: Optional[bytes] = None,
        file_obj: Optional[Union[bytes, Any]] = None,
        user_id: int = 0,
        preview: bool = False,
        background: Optional[bool] = None
    ) -> Union[Dataset, dict]:
        """
        Registers an uploaded file in the registry under 'Uploading', dry-runs validations under 'Validating',
        and inserts CrimeEvent, Criminal, Victim, and CrimeParticipation records in transactional batches under 'Importing'.
        If any validation fails, rolls back completely and marks dataset as 'Failed'.
        """
        # Validate File Extension
        if not (file_name.lower().endswith(".csv") or file_name.lower().endswith(".xlsx") or file_name.lower().endswith(".xls")):
            raise ValueError("Unsupported file format. Only CSV and Excel (.xlsx, .xls) files are supported.")

        # If file_bytes is None, we need to extract from file_obj if possible
        if file_bytes is None and file_obj is not None:
            if isinstance(file_obj, bytes):
                file_bytes = file_obj
                file_obj = None

        # Validate Empty File size
        if file_bytes is not None and len(file_bytes) == 0:
            raise ValueError("File is empty. No data found.")

        # Create temporary storage path to preview or process
        from backend.core.config import ROOT_DIR
        upload_dir = os.path.join(ROOT_DIR, "datasets", "uploaded")
        os.makedirs(upload_dir, exist_ok=True)
        storage_filename = f"temp_preview_{int(datetime.utcnow().timestamp())}_{file_name}"
        storage_path = os.path.join(upload_dir, storage_filename)
        
        # If preview mode, validate and return BulkUploadSummary without writing to DB or files permanently
        if preview:
            try:
                if file_bytes is not None:
                    with open(storage_path, "wb") as f:
                        f.write(file_bytes)
                elif file_obj is not None:
                    with open(storage_path, "wb") as f:
                        import shutil
                        shutil.copyfileobj(file_obj, f)
                
                # Check empty file size on disk
                if os.path.exists(storage_path) and os.path.getsize(storage_path) == 0:
                    raise ValueError("File is empty. No data found.")
                
                # Run preview
                return self.run_preview(storage_path, file_name)
            finally:
                if os.path.exists(storage_path):
                    try:
                        os.remove(storage_path)
                    except Exception:
                        pass

        # Validate Duplicate Filename (excluding failed and archived ones)
        duplicate_check = self.db.query(Dataset).filter(
            Dataset.original_filename == file_name,
            Dataset.status != "Failed",
            Dataset.status != "Archived"
        ).first()
        if duplicate_check:
            raise ValueError(f"A dataset with filename '{file_name}' has already been uploaded.")

        # 1. Create dataset record under 'Uploading' status
        db_dataset = Dataset(
            name=f"dataset_{int(datetime.utcnow().timestamp())}",
            original_filename=file_name,
            display_name=display_name,
            description=description,
            source_type="CSV" if file_name.lower().endswith(".csv") else "Excel",
            row_count=0,
            column_count=0,
            file_size=len(file_bytes) if file_bytes else 0,
            status="Uploading",
            upload_status="Uploading",
            storage_path=None,
            is_active=False
        )
        self.db.add(db_dataset)
        self.db.commit()
        self.db.refresh(db_dataset)

        try:
            # Write file to storage
            real_storage_filename = f"{db_dataset.id}_{file_name}"
            real_storage_path = os.path.join(upload_dir, real_storage_filename)
            
            if file_bytes is not None:
                with open(real_storage_path, "wb") as f:
                    f.write(file_bytes)
                db_dataset.file_size = len(file_bytes)
            elif file_obj is not None:
                with open(real_storage_path, "wb") as f:
                    import shutil
                    shutil.copyfileobj(file_obj, f)
                db_dataset.file_size = os.path.getsize(real_storage_path)
                
            db_dataset.storage_path = real_storage_path
            db_dataset.upload_status = "Completed"
            self.db.commit()

            # Detect schema type from headers
            import pandas as pd
            if real_storage_path.lower().endswith(".csv"):
                df_headers = pd.read_csv(real_storage_path, nrows=0)
            else:
                df_headers = pd.read_excel(real_storage_path, nrows=0, engine="openpyxl")
            
            db_dataset.column_count = len(df_headers.columns)
            self.db.commit()
            
            from backend.services.fir_import_service import FIRImportService
            fir_importer = FIRImportService(self.db)
            schema_type = fir_importer.detect_schema_type(df_headers.columns)
            db_dataset.schema_type = schema_type
            self.db.commit()

            # Check if running in background
            is_test = False
            try:
                bind = self.db.bind
                if bind and hasattr(bind, "url") and bind.url:
                    is_test = bind.url.drivername == "sqlite" and (bind.url.database == ":memory:" or not bind.url.database)
            except Exception:
                pass
                
            is_large = db_dataset.file_size > 500_000
            run_bg = background if background is not None else (is_large and not is_test)
            
            if run_bg:
                # Background task runs in a separate thread
                import threading
                threading.Thread(
                    target=run_dataset_import_bg,
                    args=(db_dataset.id, user_id, real_storage_path, schema_type),
                    daemon=True
                ).start()
                return db_dataset
            else:
                # Run synchronously
                self.execute_import(db_dataset.id, user_id, real_storage_path, schema_type, self.db)
                self.db.refresh(db_dataset)
                return db_dataset

        except Exception as file_err:
            self.db.rollback()
            db_dataset.status = "Failed"
            db_dataset.upload_status = "Failed"
            
            summary = {
                "total_rows": 0,
                "successful_imports": 0,
                "failed_imports": 1,
                "skipped_imports": 0,
                "errors": [{"row": "unknown", "error": str(file_err)}]
            }
            db_dataset.import_summary = json.dumps(summary)
            self.db.commit()
            logger.error(f"Dataset import failed: {file_err}", exc_info=True)
            raise ValueError(f"Failed to process dataset file: {str(file_err)}")

    def run_preview(self, storage_path: str, file_name: str) -> dict:
        import numpy as np
        import re
        
        # 1. Count total rows
        total_rows = 0
        if storage_path.lower().endswith(".csv"):
            with open(storage_path, "r", encoding="utf-8", errors="ignore") as f:
                for _ in f:
                    total_rows += 1
            total_rows = max(1, total_rows - 1)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(storage_path, read_only=True)
            sheet = wb.active
            total_rows = max(1, (sheet.max_row or 1) - 1)
            wb.close()

        # 2. Parse first 10 rows
        import pandas as pd
        if storage_path.lower().endswith(".csv"):
            df = pd.read_csv(storage_path, nrows=10)
        else:
            df = pd.read_excel(storage_path, nrows=10, engine="openpyxl")
        df = df.replace({np.nan: None})
        
        from backend.services.fir_import_service import FIRImportService
        fir_importer = FIRImportService(self.db)
        schema_type = fir_importer.detect_schema_type(df.columns)

        ALIAS_MAP = {
            "case_category": "case_category", "category": "case_category",
            "gravity_offence": "gravity_offence", "gravity": "gravity_offence",
            "case_status": "case_status", "status": "case_status",
            "state": "state", "district": "district", "court": "court",
            "unit_type": "unit_type", "unit": "unit", "police_station": "unit",
            "officer_kgid": "officer_kgid", "kgid": "officer_kgid",
            "officer_name": "officer_name", "officer_rank": "officer_rank",
            "rank": "officer_rank", "officer_designation": "officer_designation",
            "designation": "officer_designation", "officer_dob": "officer_dob",
            "officer_gender": "officer_gender", "officer_blood_group": "officer_blood_group",
            "officer_physically_challenged": "officer_physically_challenged",
            "officer_appointment_date": "officer_appointment_date",
            "crime_no": "crime_no", "crimeno": "crime_no", "case_no": "case_no", "caseno": "case_no",
            "registered_date": "registered_date", "crime_registered_date": "registered_date",
            "brief_facts": "brief_facts", "incident_from_date": "incident_from_date",
            "incident_to_date": "incident_to_date", "info_received_date": "info_received_date",
            "latitude": "latitude", "lat": "latitude", "longitude": "longitude", "lng": "longitude", "lon": "longitude",
            "occurrence_brief_facts": "occurrence_brief_facts", "crime_group_name": "crime_group_name",
            "crime_head_name": "crime_head_name", "act_code": "act_code", "act_description": "act_description",
            "short_name": "short_name", "section_code": "section_code", "section_description": "section_description",
            "act_order": "act_order", "section_order": "section_order", "complainant_name": "complainant_name",
            "complainant_age": "complainant_age", "complainant_occupation": "complainant_occupation",
            "complainant_religion": "complainant_religion", "complainant_caste": "complainant_caste",
            "complainant_gender": "complainant_gender", "victim_name": "victim_name", "victim_age": "victim_age",
            "victim_gender": "victim_gender", "victim_police": "victim_police", "accused_name": "accused_name",
            "accused_age": "accused_age", "accused_gender": "accused_gender", "accused_person_id": "accused_person_id",
            "arrest_type": "arrest_type", "arrest_date": "arrest_date", "arrest_state": "arrest_state",
            "arrest_district": "arrest_district", "arrest_station": "arrest_station", "arrest_io_kgid": "arrest_io_kgid",
            "arrest_court": "arrest_court", "arrest_primary_accused_name": "arrest_primary_accused_name",
            "arrest_joint_accused_names": "arrest_joint_accused_names", "chargesheet_date": "chargesheet_date",
            "chargesheet_type": "chargesheet_type", "chargesheet_officer_kgid": "chargesheet_officer_kgid"
        } if schema_type == "fir_normalized" else {
            "fir_id": "fir_id", "firid": "fir_id", "fir_no": "fir_id", "fir_number": "fir_id", "firno": "fir_id",
            "case_id": "fir_id", "case_number": "fir_id", "crime_type": "crime_type", "crimetype": "crime_type",
            "type_of_crime": "crime_type", "offence_type": "crime_type", "offencetype": "crime_type",
            "offense_type": "crime_type", "type": "crime_type", "crime": "crime_type", "crime_category": "crime_category",
            "crimecategory": "crime_category", "category": "crime_category", "district": "district", "dist": "district",
            "district_name": "district", "districtname": "district", "police_station": "police_station",
            "policestation": "police_station", "ps": "police_station", "ps_name": "police_station", "station": "police_station",
            "station_name": "police_station", "stationname": "police_station", "police_station_name": "police_station",
            "policestationname": "police_station", "crime_date": "crime_date", "crimedate": "crime_date", "date": "crime_date",
            "date_of_crime": "crime_date", "incident_date": "crime_date", "incidentdate": "crime_date", "offence_date": "crime_date",
            "crime_time": "crime_time", "crimetime": "crime_time", "time": "crime_time", "time_of_crime": "crime_time",
            "incident_time": "crime_time", "latitude": "latitude", "lat": "latitude", "longitude": "longitude",
            "lng": "longitude", "lon": "longitude", "long": "longitude", "victim_age": "victim_age", "victimage": "victim_age",
            "victim_s_age": "victim_age", "age_of_victim": "victim_age", "accused_age": "accused_age", "accusedage": "accused_age",
            "accused_s_age": "accused_age", "age_of_accused": "age_of_accused", "criminal_age": "accused_age", "gender": "gender",
            "sex": "gender", "victim_gender": "gender", "occupation": "occupation", "victim_occupation": "occupation",
            "severity": "severity", "crime_severity": "severity", "crimeseverity": "severity", "seriousness": "severity",
            "repeat_offender": "repeat_offender", "repeatoffender": "repeat_offender", "recidivist": "repeat_offender",
            "repeat": "repeat_offender", "status": "status", "case_status": "status", "casestatus": "status",
            "crime_status": "status"
        }

        def normalize_header(header: str) -> str:
            h = str(header).strip().lower().lstrip('\ufeff')
            h = h.replace(" ", "_").replace("-", "_").replace(".", "_")
            h = re.sub(r'[^a-z0-9_]', '', h)
            return re.sub(r'_+', '_', h).strip('_')

        df.columns = [normalize_header(c) for c in df.columns]
        
        # Check required columns for legacy schema
        if schema_type != "fir_normalized":
            required_cols = {"district", "police_station", "crime_date", "crime_type"}
            actual_cols = {ALIAS_MAP.get(c, c) for c in df.columns}
            missing_cols = required_cols - actual_cols
            if missing_cols:
                import pandas as pd
                if storage_path.lower().endswith(".csv"):
                    df_raw = pd.read_csv(storage_path, nrows=0)
                else:
                    df_raw = pd.read_excel(storage_path, nrows=0, engine="openpyxl")
                raise ValueError(
                    f"Missing required columns after header normalization: {sorted(missing_cols)}. "
                    f"Columns found in file: {sorted(df_raw.columns)}"
                )

        rename_dict = {col: ALIAS_MAP[col] for col in df.columns if col in ALIAS_MAP}
        df = df.rename(columns=rename_dict)
        preview_rows = df.to_dict(orient="records")

        # 3. Validate rows in chunks
        total_errors = []
        rows_validated = 0
        for chunk_rows in self.stream_rows(storage_path, schema_type, chunk_size=5000):
            normalized_chunk_rows = []
            for r in chunk_rows:
                normalized_r = {}
                for k, v in r.items():
                    normalized_k = normalize_header(k)
                    if normalized_k in ALIAS_MAP:
                        normalized_r[ALIAS_MAP[normalized_k]] = v
                    else:
                        normalized_r[normalized_k] = v
                normalized_chunk_rows.append(normalized_r)
            if schema_type == "fir_normalized":
                chunk_errors = fir_importer.validate_rows(normalized_chunk_rows)
            else:
                chunk_errors = self.validate_legacy_rows(self.db, normalized_chunk_rows, start_row_num=rows_validated + 2)
            if chunk_errors:
                total_errors.extend(chunk_errors[:100])
            rows_validated += len(chunk_rows)

        formatted_errors = []
        for idx, e in enumerate(total_errors[:100]):
            row_match = re.search(r"Row (\d+):", e)
            row_num = int(row_match.group(1)) if row_match else (idx + 2)
            formatted_errors.append({
                "row": row_num,
                "errors": {"validation": e},
                "raw_data": {}
            })

        return {
            "schema_type": schema_type,
            "total_rows": total_rows,
            "valid_count": total_rows - len(total_errors),
            "invalid_count": len(total_errors),
            "errors": formatted_errors,
            "preview": preview_rows
        }

    def stream_rows(self, storage_path: str, schema_type: str, chunk_size: int = 5000):
        import numpy as np
        if storage_path.lower().endswith(".csv"):
            for chunk in pd.read_csv(storage_path, chunksize=chunk_size):
                chunk = chunk.replace({np.nan: None})
                yield chunk.to_dict(orient="records")
        else:
            import openpyxl
            wb = openpyxl.load_workbook(storage_path, read_only=True, data_only=True)
            sheet = wb.active
            
            headers = []
            rows = []
            for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if r_idx == 0:
                    headers = [str(cell).strip() if cell is not None else f"col_{i}" for i, cell in enumerate(row)]
                else:
                    row_dict = {}
                    for h, cell in zip(headers, row):
                        row_dict[h] = cell
                    rows.append(row_dict)
                    if len(rows) >= chunk_size:
                        yield rows
                        rows = []
            if rows:
                yield rows
            wb.close()

    def validate_legacy_rows(self, db: Session, rows: list[dict], start_row_num: int) -> list[str]:
        from backend.models.location import Location
        from backend.models.police_station import PoliceStation
        
        locations = db.query(Location).all()
        district_to_loc_id = {loc.district: loc.id for loc in locations}
        stations = db.query(PoliceStation).all()
        station_to_station_id = {s.station_name: s.id for s in stations}
        
        errors = []
        for idx, row in enumerate(rows):
            row_num = start_row_num + idx
            
            missing_fields = []
            for field in ["district", "police_station", "crime_date", "crime_type"]:
                val = row.get(field)
                if val is None or (isinstance(val, str) and val.strip() == ""):
                    missing_fields.append(field)
            if missing_fields:
                errors.append(f"Row {row_num}: Missing required fields: {', '.join(missing_fields)}")
            else:
                dist = str(row["district"]).strip()
                ps = str(row["police_station"]).strip()
                if dist not in district_to_loc_id:
                    errors.append(f"Row {row_num}: District '{dist}' not found in location master data.")
                if ps not in station_to_station_id:
                    errors.append(f"Row {row_num}: Police station '{ps}' not found in master data.")
                
                date_str = str(row["crime_date"]).strip()
                try:
                    if "/" in date_str:
                        datetime.strptime(date_str, "%d/%m/%Y").date()
                    elif "-" in date_str and len(date_str.split("-")[0]) == 2:
                        datetime.strptime(date_str, "%d-%m-%Y").date()
                    else:
                        datetime.strptime(date_str[:10], "%Y-%m-%d").date()
                except ValueError:
                    errors.append(f"Row {row_num}: Invalid date format: {date_str}.")
        return errors

    def execute_import(
        self,
        dataset_id: int,
        user_id: int,
        storage_path: str,
        schema_type: str,
        db: Session,
        db_progress: Optional[Session] = None
    ):
        if db_progress is None:
            db_progress = db
            
        db_dataset_progress = db_progress.query(Dataset).filter(Dataset.id == dataset_id).first()
        db_dataset = db.query(Dataset).filter(Dataset.id == dataset_id).first()
        
        db_dataset_progress.status = "Validating"
        db_progress.commit()
        
        try:
            total_rows = 0
            if storage_path.lower().endswith(".csv"):
                with open(storage_path, "r", encoding="utf-8", errors="ignore") as f:
                    for _ in f:
                        total_rows += 1
                total_rows = max(1, total_rows - 1)
            else:
                import openpyxl
                wb = openpyxl.load_workbook(storage_path, read_only=True)
                sheet = wb.active
                total_rows = max(1, (sheet.max_row or 1) - 1)
                wb.close()
                
            if schema_type == "fir_normalized":
                from backend.services.fir_import_service import FIRImportService
                fir_importer = FIRImportService(db)
                ALIAS_MAP = {
                    "case_category": "case_category", "category": "case_category",
                    "gravity_offence": "gravity_offence", "gravity": "gravity_offence",
                    "case_status": "case_status", "status": "case_status",
                    "state": "state", "district": "district", "court": "court",
                    "unit_type": "unit_type", "unit": "unit", "police_station": "unit",
                    "officer_kgid": "officer_kgid", "kgid": "officer_kgid",
                    "officer_name": "officer_name", "officer_rank": "officer_rank",
                    "rank": "officer_rank", "officer_designation": "officer_designation",
                    "designation": "officer_designation", "officer_dob": "officer_dob",
                    "officer_gender": "officer_gender", "officer_blood_group": "officer_blood_group",
                    "officer_physically_challenged": "officer_physically_challenged",
                    "officer_appointment_date": "officer_appointment_date",
                    "crime_no": "crime_no", "crimeno": "crime_no", "case_no": "case_no", "caseno": "case_no",
                    "registered_date": "registered_date", "crime_registered_date": "registered_date",
                    "brief_facts": "brief_facts", "incident_from_date": "incident_from_date",
                    "incident_to_date": "incident_to_date", "info_received_date": "info_received_date",
                    "latitude": "latitude", "lat": "latitude", "longitude": "longitude", "lng": "longitude", "lon": "longitude",
                    "occurrence_brief_facts": "occurrence_brief_facts", "crime_group_name": "crime_group_name",
                    "crime_head_name": "crime_head_name", "act_code": "act_code", "act_description": "act_description",
                    "short_name": "short_name", "section_code": "section_code", "section_description": "section_description",
                    "act_order": "act_order", "section_order": "section_order", "complainant_name": "complainant_name",
                    "complainant_age": "complainant_age", "complainant_occupation": "complainant_occupation",
                    "complainant_religion": "complainant_religion", "complainant_caste": "complainant_caste",
                    "complainant_gender": "complainant_gender", "victim_name": "victim_name", "victim_age": "victim_age",
                    "victim_gender": "victim_gender", "victim_police": "victim_police", "accused_name": "accused_name",
                    "accused_age": "accused_age", "accused_gender": "accused_gender", "accused_person_id": "accused_person_id",
                    "arrest_type": "arrest_type", "arrest_date": "arrest_date", "arrest_state": "arrest_state",
                    "arrest_district": "arrest_district", "arrest_station": "arrest_station", "arrest_io_kgid": "arrest_io_kgid",
                    "arrest_court": "arrest_court", "arrest_primary_accused_name": "arrest_primary_accused_name",
                    "arrest_joint_accused_names": "arrest_joint_accused_names", "chargesheet_date": "chargesheet_date",
                    "chargesheet_type": "chargesheet_type", "chargesheet_officer_kgid": "chargesheet_officer_kgid"
                }
            else:
                ALIAS_MAP = {
                    "fir_id": "fir_id", "firid": "fir_id", "fir_no": "fir_id", "fir_number": "fir_id", "firno": "fir_id",
                    "case_id": "fir_id", "case_number": "fir_id", "crime_type": "crime_type", "crimetype": "crime_type",
                    "type_of_crime": "crime_type", "offence_type": "crime_type", "offencetype": "crime_type",
                    "offense_type": "crime_type", "type": "crime_type", "crime": "crime_type", "crime_category": "crime_category",
                    "crimecategory": "crime_category", "category": "crime_category", "district": "district", "dist": "district",
                    "district_name": "district", "districtname": "district", "police_station": "police_station",
                    "policestation": "police_station", "ps": "police_station", "ps_name": "police_station", "station": "police_station",
                    "station_name": "police_station", "stationname": "police_station", "police_station_name": "police_station",
                    "policestationname": "police_station", "crime_date": "crime_date", "crimedate": "crime_date", "date": "crime_date",
                    "date_of_crime": "crime_date", "incident_date": "crime_date", "incidentdate": "crime_date", "offence_date": "crime_date",
                    "crime_time": "crime_time", "crimetime": "crime_time", "time": "crime_time", "time_of_crime": "crime_time",
                    "incident_time": "crime_time", "latitude": "latitude", "lat": "latitude", "longitude": "longitude",
                    "lng": "longitude", "lon": "longitude", "long": "longitude", "victim_age": "victim_age", "victimage": "victim_age",
                    "victim_s_age": "victim_age", "age_of_victim": "victim_age", "accused_age": "accused_age", "accusedage": "accused_age",
                    "accused_s_age": "accused_age", "age_of_accused": "age_of_accused", "criminal_age": "accused_age", "gender": "gender",
                    "sex": "gender", "victim_gender": "gender", "occupation": "occupation", "victim_occupation": "occupation",
                    "severity": "severity", "crime_severity": "severity", "crimeseverity": "severity", "seriousness": "severity",
                    "repeat_offender": "repeat_offender", "repeatoffender": "repeat_offender", "recidivist": "repeat_offender",
                    "repeat": "repeat_offender", "status": "status", "case_status": "status", "casestatus": "status",
                    "crime_status": "status"
                }

            def normalize_header(header: str) -> str:
                h = str(header).strip().lower().lstrip('\ufeff')
                h = h.replace(" ", "_").replace("-", "_").replace(".", "_")
                h = re.sub(r'[^a-z0-9_]', '', h)
                return re.sub(r'_+', '_', h).strip('_')

            # Check required columns for legacy schema
            if schema_type != "fir_normalized":
                import pandas as pd
                if storage_path.lower().endswith(".csv"):
                    df_raw = pd.read_csv(storage_path, nrows=0)
                else:
                    df_raw = pd.read_excel(storage_path, nrows=0, engine="openpyxl")
                actual_cols = {ALIAS_MAP.get(normalize_header(c), normalize_header(c)) for c in df_raw.columns}
                required_cols = {"district", "police_station", "crime_date", "crime_type"}
                missing_cols = required_cols - actual_cols
                if missing_cols:
                    raise ValueError(
                        f"Missing required columns after header normalization: {sorted(missing_cols)}. "
                        f"Columns found in file: {sorted(df_raw.columns)}"
                    )

            # 2. Validation and Duplicate Check Loop
            total_errors = []
            rows_validated = 0
            for chunk_rows in self.stream_rows(storage_path, schema_type, chunk_size=5000):
                normalized_chunk_rows = []
                for r in chunk_rows:
                    normalized_r = {}
                    for k, v in r.items():
                        normalized_k = normalize_header(k)
                        if normalized_k in ALIAS_MAP:
                            normalized_r[ALIAS_MAP[normalized_k]] = v
                        else:
                            normalized_r[normalized_k] = v
                    normalized_chunk_rows.append(normalized_r)

                if schema_type == "fir_normalized":
                    chunk_errors = fir_importer.validate_rows(normalized_chunk_rows)
                else:
                    chunk_errors = self.validate_legacy_rows(db, normalized_chunk_rows, start_row_num=rows_validated + 2)
                    
                if chunk_errors:
                    total_errors.extend(chunk_errors[:100])
                    
                rows_validated += len(chunk_rows)
                
                progress = int((rows_validated / total_rows) * 50)
                db_dataset_progress.import_summary = json.dumps({
                    "progress": progress,
                    "status": "Validating",
                    "processed_rows": rows_validated,
                    "total_rows": total_rows
                })
                db_progress.commit()
                
            if total_errors:
                raise ValueError("Validation errors: " + " | ".join(total_errors[:50]))
                
            # 3. Transition to Ingesting/Importing
            db_dataset_progress.status = "Importing"
            db_progress.commit()
            
            # 4. Ingestion Loop
            rows_imported = 0
            summary_reports = []
            
            if schema_type == "fir_normalized":
                fir_importer.preload_caches()
            else:
                from backend.models.location import Location
                from backend.models.police_station import PoliceStation
                from backend.models.crime import CrimeEvent
                from backend.models.criminal import Criminal
                from backend.models.victim import Victim
                from backend.models.crime_participation import CrimeParticipation
                
                locations = db.query(Location).all()
                district_to_loc_id = {loc.district: loc.id for loc in locations}
                stations = db.query(PoliceStation).all()
                station_to_station_id = {s.station_name: s.id for s in stations}
                
                def clean_ps_name(name):
                    name = str(name).strip()
                    name = re.sub(r'\s+', ' ', name)
                    if name.lower().endswith(" ps"):
                        return f"{name[:-3].strip().title()} PS"
                    elif name.lower().endswith("ps"):
                        return f"{name[:-2].strip().title()} PS"
                    return name.title()
                    
                def get_loc_id(dist_name):
                    dist_name = str(dist_name).strip().title()
                    if dist_name in district_to_loc_id:
                        return district_to_loc_id[dist_name]
                    new_loc = Location(state="Karnataka", district=dist_name)
                    db.add(new_loc)
                    db.flush()
                    district_to_loc_id[dist_name] = new_loc.id
                    return new_loc.id

                def get_ps_id(ps_name, dist_name):
                    ps_name = clean_ps_name(ps_name)
                    if ps_name in station_to_station_id:
                        return station_to_station_id[ps_name]
                    l_id = get_loc_id(dist_name)
                    new_ps = PoliceStation(
                        station_name=ps_name,
                        district=dist_name,
                        location_id=l_id,
                        beat="Beat A",
                        officer_count=10,
                        vehicle_count=2,
                        equipment_count=5,
                        capacity=20,
                        availability="active"
                    )
                    db.add(new_ps)
                    db.flush()
                    station_to_station_id[ps_name] = new_ps.id
                    return new_ps.id

                def parse_date_robust(date_val):
                    if not date_val or str(date_val).strip().lower() in ["nan", "none", "null", ""]:
                        return None
                    if isinstance(date_val, (date, datetime)):
                        return date_val.date() if isinstance(date_val, datetime) else date_val
                    d_str = str(date_val).strip()
                    if "t" in d_str.lower():
                        d_str = re.split(r'[tT]', d_str)[0]
                    elif " " in d_str:
                        d_str = d_str.split(" ")[0]
                    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                        try:
                            return datetime.strptime(d_str, fmt).date()
                        except ValueError:
                            pass
                    return None

                def parse_time_robust(time_val):
                    from datetime import time as dt_time
                    if not time_val or str(time_val).strip().lower() in ["nan", "none", "null", ""]:
                        return None
                    if isinstance(time_val, datetime):
                        return time_val.time()
                    if isinstance(time_val, dt_time):
                        return time_val
                    t_str = str(time_val).strip()
                    if "t" in t_str.lower():
                        t_str = re.split(r'[tT]', t_str)[1]
                    elif " " in t_str:
                        parts = t_str.split(" ")
                        if len(parts) > 1 and ":" in parts[1]:
                            t_str = parts[1]
                    t_str = t_str[:8]
                    for fmt in ("%H:%M:%S", "%H:%M"):
                        try:
                            return datetime.strptime(t_str, fmt).time()
                        except ValueError:
                            pass
                    return None

            legacy_criminal_cache = {}
            for chunk_rows in self.stream_rows(storage_path, schema_type, chunk_size=5000):
                normalized_chunk_rows = []
                for r in chunk_rows:
                    normalized_r = {}
                    for k, v in r.items():
                        normalized_k = normalize_header(k)
                        if normalized_k in ALIAS_MAP:
                            normalized_r[ALIAS_MAP[normalized_k]] = v
                        else:
                            normalized_r[normalized_k] = v
                    normalized_chunk_rows.append(normalized_r)
                    
                if schema_type == "fir_normalized":
                    report = fir_importer.import_normalized_dataset(normalized_chunk_rows, dataset_id, user_id=None, commit=False)
                    summary_reports.append(report)
                    rows_imported += report.get("cases_inserted", 0)
                else:
                    successful_imports = 0
                    
                    # Group chunk rows by unique case ID
                    grouped_legacy_cases = {}
                    for idx, r in enumerate(normalized_chunk_rows):
                        case_key = r.get("fir_id") or r.get("crime_no") or f"gen_{rows_imported + idx}"
                        case_key = str(case_key).strip()
                        if case_key not in grouped_legacy_cases:
                            grouped_legacy_cases[case_key] = []
                        grouped_legacy_cases[case_key].append(r)
                        
                    for case_key, case_rows in grouped_legacy_cases.items():
                        first_row = case_rows[0]
                        dist = str(first_row.get("district", "")).strip()
                        ps = str(first_row.get("police_station", "")).strip()
                        
                        loc_id = get_loc_id(dist)
                        ps_id = get_ps_id(ps, dist)
                        
                        date_obj = parse_date_robust(first_row.get("crime_date")) or date.today()
                        time_obj = parse_time_robust(first_row.get("crime_time"))
                        
                        crime = CrimeEvent(
                            crime_type=str(first_row.get("crime_type", "General Crime")),
                            crime_category=str(first_row.get("crime_category") or first_row.get("crime_type") or "General"),
                            crime_subcategory=first_row.get("crime_subcategory") or f"{first_row.get('crime_type', 'General')} - Subcategory",
                            description=first_row.get("description") or first_row.get("brief_facts") or f"{first_row.get('crime_type')} incident registered under {case_key}",
                            severity=float(first_row["severity"]) if first_row.get("severity") else 1.0,
                            status=str(first_row.get("status", "reported")),
                            crime_date=date_obj,
                            crime_time=time_obj,
                            location_id=loc_id,
                            police_station_id=ps_id,
                            dataset_id=dataset_id,
                            victim_count=0,
                            accused_count=0
                        )
                        db.add(crime)
                        db.flush()
                        
                        seen_accused = set()
                        seen_victims = set()
                        
                        for r in case_rows:
                            acc_name = r.get("accused_name") or r.get("accused")
                            if not acc_name or str(acc_name).strip().lower() in ["nan", "none", "null", ""]:
                                # Fall back to Unknown Accused if some other attributes are present
                                if r.get("accused_age") or r.get("gender") or r.get("accused_gender") or r.get("repeat_offender"):
                                    acc_name = "Unknown Accused"
                                else:
                                    acc_name = None
                            
                            if acc_name:
                                seen_accused.add(str(acc_name).strip().title())
                                
                            joint_str = r.get("arrest_joint_accused_names")
                            if joint_str and str(joint_str).strip().lower() not in ["nan", "none", "null", ""]:
                                for name in str(joint_str).split(","):
                                    if name.strip():
                                        seen_accused.add(name.strip().title())
                                        
                            vic_name = r.get("victim_name") or r.get("victim")
                            if not vic_name or str(vic_name).strip().lower() in ["nan", "none", "null", ""]:
                                if r.get("victim_age") or r.get("gender") or r.get("victim_gender"):
                                    vic_name = "Unknown Victim"
                                else:
                                    vic_name = None
                                    
                            if vic_name:
                                seen_victims.add(str(vic_name).strip().title())
                                
                        crime.accused_count = len(seen_accused)
                        crime.victim_count = len(seen_victims)
                        db.flush()
                        
                        # Create Criminals and Participations
                        for acc_name in seen_accused:
                            acc_row = next((r for r in case_rows if str(r.get("accused_name") or r.get("accused") or "").strip().title() == acc_name), None)
                            if not acc_row and acc_name == "Unknown Accused":
                                acc_row = next((r for r in case_rows if r.get("accused_age") or r.get("gender") or r.get("accused_gender")), first_row)
                                
                            gender = str(acc_row.get("gender") or acc_row.get("accused_gender") or "Male").strip().title() if acc_row else "Male"
                            age = float(acc_row["accused_age"]) if acc_row and acc_row.get("accused_age") is not None else None
                            caste = str(acc_row.get("caste")).strip() if acc_row and acc_row.get("caste") else None
                            occ = str(acc_row.get("occupation")).strip() if acc_row and acc_row.get("occupation") else None
                            risk = 0.85 if acc_row and str(acc_row.get("repeat_offender")) == "1" else 0.15
                            
                            is_unknown = (acc_name.lower() == "unknown accused")
                            cache_key = (acc_name.lower(), gender, age)
                            if cache_key in legacy_criminal_cache and not is_unknown:
                                crim_id = legacy_criminal_cache[cache_key]
                            else:
                                c_rec = None
                                if not is_unknown:
                                    c_rec = db.query(Criminal).filter(
                                        Criminal.name == acc_name,
                                        Criminal.dataset_id == dataset_id
                                    ).first()
                                if not c_rec:
                                    c_rec = Criminal(
                                        name=acc_name,
                                        gender=gender,
                                        age=age,
                                        occupation=occ,
                                        caste=caste,
                                        risk_score=risk,
                                        status="accused" if first_row.get("status") != "Closed" else "convicted",
                                        dataset_id=dataset_id
                                    )
                                    db.add(c_rec)
                                    db.flush()
                                crim_id = c_rec.id
                                if not is_unknown:
                                    legacy_criminal_cache[cache_key] = crim_id
                                
                            part = CrimeParticipation(
                                crime_event_id=crime.id,
                                criminal_id=crim_id,
                                role="principal accused" if len(seen_accused) == 1 else "co-offender",
                                dataset_id=dataset_id
                            )
                            db.add(part)
                            
                        # Create Victims
                        for vic_name in seen_victims:
                            vic_row = next((r for r in case_rows if str(r.get("victim_name") or r.get("victim") or "").strip().title() == vic_name), None)
                            if not vic_row and vic_name == "Unknown Victim":
                                vic_row = next((r for r in case_rows if r.get("victim_age") or r.get("gender") or r.get("victim_gender")), first_row)
                                
                            v_gender = str(vic_row.get("gender") or vic_row.get("victim_gender") or "Male").strip().title() if vic_row else "Male"
                            v_age = float(vic_row["victim_age"]) if vic_row and vic_row.get("victim_age") is not None else None
                            v_occ = str(vic_row.get("occupation") or vic_row.get("victim_occupation", "")).strip() if vic_row else None
                            if v_occ == "":
                                v_occ = None
                                
                            legacy_vic = Victim(
                                crime_event_id=crime.id,
                                gender=v_gender,
                                age=v_age,
                                occupation=v_occ,
                                location_id=loc_id,
                                dataset_id=dataset_id
                            )
                            db.add(legacy_vic)
                            
                        successful_imports += len(case_rows)
                        
                    db.flush()
                    rows_imported += successful_imports
                    
                progress = 50 + int((rows_imported / total_rows) * 50)
                db_dataset_progress.import_summary = json.dumps({
                    "progress": progress,
                    "status": "Importing",
                    "processed_rows": rows_imported,
                    "total_rows": total_rows
                })
                db_progress.commit()
                
            db.commit()
            
            if schema_type == "fir_normalized":
                summary = {
                    "total_rows": total_rows,
                    "cases_inserted": sum(r.get("cases_inserted", 0) for r in summary_reports),
                    "victims_inserted": sum(r.get("victims_inserted", 0) for r in summary_reports),
                    "accused_inserted": sum(r.get("accused_inserted", 0) for r in summary_reports),
                    "complainants_inserted": sum(r.get("complainants_inserted", 0) for r in summary_reports),
                    "arrests_inserted": sum(r.get("arrests_inserted", 0) for r in summary_reports),
                    "chargesheets_inserted": sum(r.get("chargesheets_inserted", 0) for r in summary_reports),
                    "warnings": [w for r in summary_reports for w in r.get("warnings", [])]
                }
            else:
                summary = {
                    "total_rows": total_rows,
                    "successful_imports": rows_imported,
                    "failed_imports": 0,
                    "skipped_imports": 0,
                    "errors": []
                }
                
            db_dataset_progress.status = "Ready"
            db_dataset_progress.row_count = rows_imported
            db_dataset_progress.import_summary = json.dumps(summary)
            db_progress.commit()
            
            # Auto-activate completed dataset
            db_dataset_ref = db.query(Dataset).filter(Dataset.id == dataset_id).first()
            if db_dataset_ref:
                self.activate_dataset(db_dataset_ref.id)
                
        except Exception as e:
            db.rollback()
            db_dataset_progress.status = "Failed"
            db_dataset_progress.upload_status = "Failed"
            db_dataset_progress.import_summary = json.dumps({
                "progress": 0,
                "status": "Failed",
                "errors": [{"row": "unknown", "error": str(e)}]
            })
            db_progress.commit()
            raise e

    def get_dataset_summary(self, dataset_id: int) -> dict:
        """
        Compiles high-level statistical summary metrics for a given dataset,
        including total crimes, unique criminals, unique victims, geographic
        coverage, and date bounds.
        """
        from backend.models.crime import CrimeEvent
        from backend.models.criminal import Criminal
        from backend.models.victim import Victim
        from backend.models.location import Location
        from sqlalchemy import func

        dataset = self.db.query(Dataset).filter(Dataset.id == dataset_id).first()
        if not dataset:
            raise ValueError(f"Dataset with ID {dataset_id} not found.")

        total_crimes = self.db.query(CrimeEvent).filter(CrimeEvent.dataset_id == dataset_id).count()
        criminals = self.db.query(Criminal).filter(Criminal.dataset_id == dataset_id).count()
        victims = self.db.query(Victim).filter(Victim.dataset_id == dataset_id).count()

        # Date range bounds
        min_date, max_date = self.db.query(
            func.min(CrimeEvent.crime_date),
            func.max(CrimeEvent.crime_date)
        ).filter(CrimeEvent.dataset_id == dataset_id).first()

        # Unique districts list
        districts_query = self.db.query(Location.district).join(CrimeEvent).filter(
            CrimeEvent.dataset_id == dataset_id
        ).distinct().all()
        districts = [row[0] for row in districts_query if row[0]]

        return {
            "total_crimes": total_crimes,
            "criminals": criminals,
            "victims": victims,
            "date_range": {
                "min": min_date.isoformat() if min_date else None,
                "max": max_date.isoformat() if max_date else None
            },
            "districts": districts,
            "upload_time": dataset.upload_time,
            "file_size": dataset.file_size
        }

    def get_dataset_preview(self, dataset_id: int) -> dict:
        """
        Reads the first 20 rows and column metadata from the stored dataset file.
        """
        dataset = self.db.query(Dataset).filter(Dataset.id == dataset_id).first()
        if not dataset:
            raise ValueError(f"Dataset with ID {dataset_id} not found.")

        # Resolve path
        path = dataset.storage_path
        if not path or not os.path.exists(path):
            raise ValueError(f"Dataset storage file not found for dataset ID {dataset_id}.")

        # Load file
        if path.lower().endswith(".xlsx") or path.lower().endswith(".xls"):
            df = pd.read_excel(path, engine="openpyxl")
        else:
            df = pd.read_csv(path)

        df = df.replace({np.nan: None})

        # Columns & Types
        dtypes_map = {}
        for col, dtype in df.dtypes.items():
            dtype_str = str(dtype)
            if "int" in dtype_str:
                dtypes_map[col] = "integer"
            elif "float" in dtype_str:
                dtypes_map[col] = "float"
            elif "datetime" in dtype_str:
                dtypes_map[col] = "datetime"
            elif "bool" in dtype_str:
                dtypes_map[col] = "boolean"
            else:
                dtypes_map[col] = "string"

        return {
            "first_20_rows": df.head(20).to_dict(orient="records"),
            "total_rows": len(df),
            "total_columns": len(df.columns),
            "columns": list(df.columns),
            "data_types": dtypes_map
        }

    def get_dataset_statistics(self, dataset_id: int) -> dict:
        """
        Computes row/column count, missing values, duplicates, and numeric/categorical columns.
        """
        dataset = self.db.query(Dataset).filter(Dataset.id == dataset_id).first()
        if not dataset:
            raise ValueError(f"Dataset with ID {dataset_id} not found.")

        # Resolve path
        path = dataset.storage_path
        if not path or not os.path.exists(path):
            raise ValueError(f"Dataset storage file not found for dataset ID {dataset_id}.")

        # Load file
        if path.lower().endswith(".xlsx") or path.lower().endswith(".xls"):
            df = pd.read_excel(path, engine="openpyxl")
        else:
            df = pd.read_csv(path)

        missing_values = {col: int(val) for col, val in df.isnull().sum().items()}
        duplicate_rows = int(df.duplicated().sum())
        
        numeric_columns = df.select_dtypes(include=[np.number]).columns.tolist()
        categorical_columns = df.select_dtypes(exclude=[np.number]).columns.tolist()

        return {
            "total_rows": int(df.shape[0]),
            "total_columns": int(df.shape[1]),
            "missing_values": missing_values,
            "duplicate_rows": duplicate_rows,
            "numeric_columns": numeric_columns,
            "categorical_columns": categorical_columns
        }


def run_dataset_import_bg(dataset_id: int, user_id: int, storage_path: str, schema_type: str):
    from backend.core.database import SessionLocal
    from backend.services.dataset_service import DatasetService
    db = SessionLocal()
    try:
        service = DatasetService(db)
        service.execute_import(dataset_id, user_id, storage_path, schema_type, db, db)
    except Exception as e:
        logger.error(f"Background dataset import failed for ID {dataset_id}: {e}", exc_info=True)
    finally:
        db.close()

